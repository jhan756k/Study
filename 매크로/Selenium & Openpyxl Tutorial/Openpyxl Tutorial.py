from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.title = "Test"
ws1 = wb.create_sheet("Test1")
print(wb.sheetnames)

test1cel = ws['A4']
ws.cell(row=4, column=2, value=10)
test1cel.value = 5

for row in ws.iter_rows(min_row=1, max_col=1, max_row=4): # iter rows, columns, etc (DOCUMENTATION)
    for cell in row:
        print(cell.value)
        
wb.save('macrotest.xlsx')











































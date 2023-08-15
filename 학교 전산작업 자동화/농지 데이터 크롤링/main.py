from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import Font
from openpyxl import load_workbook
wb = load_workbook(filename = "./farmdata.xlsx")
ws = wb.active
fontStyle = Font(size = "20")

driver = webdriver.Chrome()

url = "https://www.alimi.or.kr/dataview/a/totalSaleSearchList2.do"

'''
/html/body/div/form/div/div[1]/div/div[1]/div[3]/ul[1]/li/a/span[1]
          0 이것만 변함 

다음 페이지에서는 동일함
1 ~ 20
'''

driver.get(url)
driver.maximize_window()

'''
시군구
주재배작물
면적 (m^2)
소유권변동일자
소유구분
개별공시지가
공시일자
'''

ind = 2

for i in range(1, 14):

    if i == 6 or i == 11:
        driver.find_element(By.XPATH, "/html/body/div/form/div/div[1]/div/div[1]/div[4]/ul/li[8]/a").click()
    else:
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
            (By.LINK_TEXT, str(i)))).click()

    for n in range(1, 21):

        landtype = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
        (By.XPATH, f"/html/body/div/form/div/div[1]/div/div[1]/div[3]/ul[{n}]/li/a/span[1]")))
        landtype.click()

        sigungu = driver.find_element(By.XPATH, "/html/body/div/form/div/div[1]/div/div[3]/dt[1]/span").text

        crop = driver.find_element(By.XPATH, "/html/body/div/form/div/div[1]/div/div[3]/dd[5]").text[8:]

        area = driver.find_element(By.XPATH, "/html/body/div/form/div/div[1]/div/div[3]/dd[6]").text[8:]

        soyudate = driver.find_element(By.XPATH, "/html/body/div/form/div/div[1]/div/div[3]/dd[10]").text[10:]

        whosoyu = driver.find_element(By.XPATH, "/html/body/div/form/div/div[1]/div/div[3]/dd[11]").text[7:]

        gongsijilja = driver.find_element(By.XPATH, "/html/body/div/form/div/div[1]/div/div[3]/dd[15]").text[9:]

        gongsidate = driver.find_element(By.XPATH, "/html/body/div/form/div/div[1]/div/div[3]/dd[17]").text[7:]

        ws.cell(row=ind, column=1, value = landtype.text).font = fontStyle    
        ws.cell(row=ind, column=2).value = sigungu
        ws.cell(row=ind, column=3).value = crop
        ws.cell(row=ind, column=4).value = area
        ws.cell(row=ind, column=5).value = soyudate
        ws.cell(row=ind, column=6).value = whosoyu
        ws.cell(row=ind, column=7).value = gongsijilja
        ws.cell(row=ind, column=8).value = gongsidate

        ind += 1

wb.save("result.xlsx")
driver.quit()

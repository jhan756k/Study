import openpyxl

wb = openpyxl.load_workbook('name.xlsx')
name27 = wb.get_sheet_by_name(r'2학년(27기)')
name26 = wb.get_sheet_by_name(r'3학년(26기)')

samuchim = "김강민 김태희 신이현 이창목 라연준 김정우 김규진 류의현 유강은 이다은 정민서 정수찬 배강민 이윤호 하재균 서지원 김헌 오승찬 박지영 정서희​ 인선조 이준우"

sah = "노준영 김주현"

sisang = "정유찬A 엄세아" 

ule = "김솔 민규리 이준우 인선조 최시현"

gisu = "최유하 최준원 김동건 유호권"

daechi = "김택헌 김지민 정유찬B 강호성 김유신 김성주 김환희 강동혁 서인성"

bangsong =  "안보영 최수민 오승찬 정민교 조재아​"

jiwh = "오민준" 

docnip = "황규빈 전서영"

names = [samuchim, sah, sisang, ule, gisu, daechi, bangsong, jiwh, docnip]

ans = ""

breaks = True
breaks1 = True

print()

for group in names:
    namelist = group.split()
    for ppl in namelist:

        for i in range(3, 14):
            if not breaks:
                breaks = True
                break
            for j in range(3, 19):
                if name27.cell(row=j, column=i).value == ppl:
                    ans += (ppl + " 2-" + str(i-2) + " ")
                    breaks = False
                    break
        else:
            for i in range(3, 14):
                if not breaks1:
                    breaks1 = True
                    break
                for j in range(4, 23):
                    if name26.cell(row=j, column=i).value == ppl:
                        ans += (ppl + " 3-" + str(i-2) + " ")
                        breaks1 = False
                        break

print(ans)

# 사무침: 김강민 2-2 김태희 2-6 신이현 2-1 이창목 2-3 라연준 2-1 김정우 2-2 김규진 2-6 류의현 2-7 유강은 2-10 이다은 2-9 정민서 2-1 정수찬 2-5 배강민 2-1 이윤호 2-6 하재균 2-6 서지원 2-7 김헌 2-1 오승찬 2-2 박지영 2-4 정서희 2-6 인선조 3-5 이준우 3-6 

# 사회자: 노준영 3-2 김주현 3-8 

# 시상도우미: 정유찬A 2-9 엄세아 2-7

# 의례단 명단 수정: 김솔 3-4 민규리 3-6 이준우 3-6 인선조 3-5 최시현 3-4 

# 기수단: 최유하 3-10 최준원 3-1 김동건 3-2 유호권 3-7 

# 대취타: 김택헌 2-2 김지민 2-4 정유찬B 2-4 강호성 2-3 김유신 2-7 김성주 2-7 김환희 2-6 강동혁 2-7 서인성 2-5 

# 방송부: 안보영 2-6 최수민 2-9 오승찬 2-2 정민교 3-5 조재아 3-2 

# 지휘자: 오민준 2-4 

# 독립선언서 낭독자: 황규빈 3-10 전서영 3-7
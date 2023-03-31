from openpyxl import load_workbook
f = open("d.txt", 'rt', encoding = "UTF8")
sangf = f.read()
sangf_tmp = str(sangf.replace("\n", " "))

sangsang = sangf_tmp.split(" ")



sangsanglist = []


[sangsanglist.append(x) for x in sangsang if x != ""]

load_wb = load_workbook("ans.xlsx", data_only=True)
ws = load_wb['Sheet1']

for x in range(len(sangsanglist)):
    ws.cell(x+1, 1).value = sangsanglist[x]

load_wb.save('ans.xlsx')